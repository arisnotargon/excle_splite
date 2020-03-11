import openpyxl
import base64
import os
import re
import argparse
import time


def make_us_base(temp_sheet_us, name, sid, year):
    temp_sheet_us.cell(2, 1).value = 'SRUSA'
    temp_sheet_us.cell(2, 2).value = name
    temp_sheet_us.cell(2, 7).value = year + ' PAY OVERVIEW'
    temp_sheet_us.cell(3, 2).value = sid
    temp_sheet_us.cell(4, 2).value = 'US$'
    temp_sheet_us.cell(5, 2).value = 'January'
    temp_sheet_us.cell(5, 3).value = 'February'
    temp_sheet_us.cell(5, 4).value = 'March'
    temp_sheet_us.cell(5, 5).value = 'Special Bonus'
    temp_sheet_us.cell(5, 6).value = 'April'
    temp_sheet_us.cell(5, 7).value = 'May'
    temp_sheet_us.cell(5, 8).value = 'June'
    temp_sheet_us.cell(5, 9).value = 'Summer Bonus'
    temp_sheet_us.cell(5, 10).value = 'July'
    temp_sheet_us.cell(5, 11).value = 'August'
    temp_sheet_us.cell(5, 12).value = 'September'
    temp_sheet_us.cell(5, 13).value = 'October'
    temp_sheet_us.cell(5, 14).value = 'November'
    temp_sheet_us.cell(5, 15).value = 'December'
    temp_sheet_us.cell(5, 16).value = 'Winter Bonus'
    temp_sheet_us.cell(5, 17).value = 'TOTAL'
    temp_sheet_us.cell(11, 1).value = 'Adjustment'
    temp_sheet_us.cell(12, 1).value = 'TOTAL'
    temp_sheet_us.cell(13, 1).value = 'Loan Repayment'
    temp_sheet_us.cell(14, 1).value = 'Net pay in USA'
    temp_sheet_us.cell(15, 1).value = 'Gross pay in US'
    temp_sheet_us.cell(17, 1).value = 'Paid in Japan'
    temp_sheet_us.cell(18, 1).value = 'YEN'
    temp_sheet_us.cell(19, 1).value = 'Adjustment'
    temp_sheet_us.cell(20, 1).value = 'Total (YEN)'
    for i in range(16):
        temp_sheet_us.cell(17, i + 2).value = 'YEN'


def make_jp_base(temp_sheet_us, name, sid, year):
    temp_sheet_us.cell(1, 1).value = sid
    temp_sheet_us.cell(1, 2).value = name
    temp_sheet_us.cell(1, 2).value = 'Income for ' + year + '(amount in Japanese yen)'
    temp_sheet_us.cell(3, 1).value = 'Unit: YEN'
    temp_sheet_us.cell(4, 2).value = 'Payment'
    temp_sheet_us.cell(4, 3).value = 'Deduction'
    temp_sheet_us.cell(5, 2).value = 'Gloss payment in Japan'
    temp_sheet_us.cell(5, 3).value = 'Income Tax'
    temp_sheet_us.cell(5, 4).value = 'Socail Insurance premium'
    temp_sheet_us.cell(5, 5).value = 'Inhabitant  Tax'
    temp_sheet_us.cell(5, 6).value = 'Other Deduction'
    temp_sheet_us.cell(6, 4).value = 'Welfare Pension Insurance Premium'
    temp_sheet_us.cell(6, 5).value = 'Health Insurance Premium'
    temp_sheet_us.cell(6, 6).value = 'Unemployment Insurance premium'
    temp_sheet_us.cell(7, 1).value = 'January'
    temp_sheet_us.cell(8, 1).value = 'February'
    temp_sheet_us.cell(9, 1).value = 'March'
    temp_sheet_us.cell(10, 1).value = 'Special Bonus'
    temp_sheet_us.cell(11, 1).value = 'April'
    temp_sheet_us.cell(12, 1).value = 'May'
    temp_sheet_us.cell(13, 1).value = 'June'
    temp_sheet_us.cell(14, 1).value = 'Summer Bonus'
    temp_sheet_us.cell(15, 1).value = 'July'
    temp_sheet_us.cell(16, 1).value = 'August'
    temp_sheet_us.cell(17, 1).value = 'September'
    temp_sheet_us.cell(18, 1).value = 'October'
    temp_sheet_us.cell(19, 1).value = 'November'
    temp_sheet_us.cell(20, 1).value = 'December'
    temp_sheet_us.cell(21, 1).value = 'Winter Bonus'
    temp_sheet_us.cell(22, 1).value = 'Total:'
    temp_sheet_us.merge_cells('A4:A6')
    temp_sheet_us.merge_cells('C4:H6')
    temp_sheet_us.merge_cells('D5:F5')
    temp_sheet_us.merge_cells('B5:B6')
    temp_sheet_us.merge_cells('C5:C6')
    temp_sheet_us.merge_cells('G5:G6')
    temp_sheet_us.merge_cells('H5:H6')


def xls_splite_by_name(source_file_name):
    print(source_file_name)
    dotinx = source_file_name.rfind('.')
    out_name = source_file_name[:dotinx] + '_out' + '_' + str(int(time.time()))
    if dotinx >= 0:
        out_name = out_name + source_file_name[dotinx:]
    source_wb = openpyxl.load_workbook(source_file_name)
    source_sheet = source_wb.active
    lineNo = 5
    source_data = {}
    while True:
        temp_staff_id = source_sheet.cell(lineNo, 3).value
        if temp_staff_id == None:
            break

        if not temp_staff_id in source_data:
            source_data[temp_staff_id] = {}
        source_data[temp_staff_id]['name'] = source_sheet.cell(lineNo, 7).value
        year_month = source_sheet.cell(lineNo, 9).value
        source_data[temp_staff_id][year_month] = {}
        source_data[temp_staff_id][year_month]['us'] = source_sheet.cell(lineNo, 35).value
        source_data[temp_staff_id][year_month]['jp'] = source_sheet.cell(lineNo, 59).value
        source_data[temp_staff_id][year_month]['unemployment'] = source_sheet.cell(lineNo, 65).value
        source_data[temp_staff_id][year_month]['welfare'] = source_sheet.cell(lineNo, 68).value
        source_data[temp_staff_id][year_month]['health'] = source_sheet.cell(lineNo, 66).value
        lineNo += 1

    for sid in source_data:
        temp_sheet_name = source_data[sid]['name']
        if temp_sheet_name in source_wb.sheetnames:
            temp_sheet_name = temp_sheet_name + str(sid)

        year = ''
        for k in source_data[sid].keys():
            if isinstance(k, int):
                year = str(k)[:4]

        temp_sheet_us = source_wb.create_sheet(temp_sheet_name)
        temp_sheet_jp = source_wb.create_sheet(temp_sheet_name + '(J)')

        make_us_base(temp_sheet_us, source_data[sid]['name'], sid, year)
        make_jp_base(temp_sheet_jp, source_data[sid]['name'], sid, year)
        jpe_sum = 0
        usd_sum = 0
        unemployment_sum = 0
        welfare_sum = 0
        health_sum = 0
        for k in source_data[sid].keys():
            if k != 'name':
                # print(source_data[sid][k])
                # exit()
                jpe_sum += int(source_data[sid][k]['jp'])
                usd_sum += int(source_data[sid][k]['us'])
                unemployment_sum += int(source_data[sid][k]['unemployment'])
                welfare_sum += int(source_data[sid][k]['welfare'])
                health_sum += int(source_data[sid][k]['health'])
                month = int(k) % 100
                us_col = month + 1
                jp_row = month + 6
                if month > 3:
                    us_col += 1
                    jp_row += 1

                if month > 6:
                    us_col += 1
                    jp_row += 1

                temp_sheet_us.cell(6, us_col).value = source_data[sid][k]['us']
                temp_sheet_us.cell(12, us_col).value = source_data[sid][k]['us']
                temp_sheet_us.cell(14, us_col).value = source_data[sid][k]['us']
                temp_sheet_us.cell(18, us_col).value = source_data[sid][k]['jp']
                temp_sheet_us.cell(20, us_col).value = source_data[sid][k]['jp']
                temp_sheet_jp.cell(jp_row, 2).value = source_data[sid][k]['jp']
                temp_sheet_jp.cell(jp_row, 3).value = 0
                temp_sheet_jp.cell(jp_row, 4).value = source_data[sid][k]['welfare']
                temp_sheet_jp.cell(jp_row, 5).value = source_data[sid][k]['health']
                temp_sheet_jp.cell(jp_row, 6).value = source_data[sid][k]['unemployment']
                temp_sheet_jp.cell(jp_row, 7).value = 0
                temp_sheet_jp.cell(jp_row, 8).value = \
                    int(source_data[sid][k]['welfare']) \
                    + int(source_data[sid][k]['health']) \
                    + source_data[sid][k]['unemployment']

        temp_sheet_us.cell(6, 17).value = usd_sum
        temp_sheet_us.cell(12, 17).value = usd_sum
        temp_sheet_us.cell(14, 17).value = usd_sum
        temp_sheet_us.cell(18, 17).value = jpe_sum
        temp_sheet_us.cell(20, 17).value = jpe_sum
        temp_sheet_jp.cell(22, 2).value = jpe_sum
        temp_sheet_jp.cell(22, 3).value = 0
        temp_sheet_jp.cell(22, 4).value = welfare_sum
        temp_sheet_jp.cell(22, 5).value = health_sum
        temp_sheet_jp.cell(22, 6).value = unemployment_sum
        temp_sheet_jp.cell(22, 7).value = 0
        temp_sheet_jp.cell(22, 8).value = welfare_sum + health_sum + unemployment_sum

    source_wb.save(out_name)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--source-file-name', dest='sfname', required=True,
                        help='name of source excel file')
    sfname = parser.parse_args().sfname
    # sfname = 'byname.xlsx'

    xls_splite_by_name(sfname)
