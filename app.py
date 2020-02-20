import openpyxl
import base64
import os
import re
import argparse

def xls_splite(source_file_name):
    source_wb = openpyxl.load_workbook(source_file_name)
    source_sheet = source_wb.active
    headers = source_sheet[1:4]
    tmp_data = {}

    line_no = 4
    while True:
        line_no += 1
        line_first_cell = source_sheet.cell(line_no, 1)
        if not (bool(line_first_cell.value)):
            break

        #bo=business office
        tmp_bo = source_sheet.cell(line_no,4).value
        tmp_bo_key = base64.encodebytes(bytes(tmp_bo,"UTF-8")).hex()

        if not (tmp_bo_key in tmp_data.keys()):
            tmp_data[tmp_bo_key] = {
                "name": tmp_bo,
                "data": []
            }

        tmp_data[tmp_bo_key]["data"].append(source_sheet[line_no])

    dir_name = source_file_name+'_splite'
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)

    for bo in tmp_data:
        des_file_name = dir_name + os.sep + re.sub('[\/\\\:\*\?\"\<\>\|]', "_", tmp_data[bo]['name']) + '.xlsx'
        des_wb = openpyxl.Workbook()
        des_sheet = des_wb.active
        for line in headers:
            for cell in line:
                des_sheet.cell(cell.row,cell.column).value = cell.value

        line_no = 4
        for line in tmp_data[bo]['data']:
            line_no += 1
            for cell in line:
                des_sheet.cell(line_no, cell.column).value = cell.value

        des_wb.save(des_file_name)
        des_wb.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--source-file-name', dest='sfname', required=True,
                        help='name of source excel file')
    sfname = parser.parse_args().sfname
    xls_splite(sfname)